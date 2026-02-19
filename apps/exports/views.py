"""
Export API Views
"""
import io
import csv
from django.db import models
from django.http import HttpResponse
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response

from .models import ExportTemplate, ExportHistory, ReportSchedule
from .serializers import (
    ExportTemplateSerializer, ExportHistorySerializer, ReportScheduleSerializer
)


class ExportTemplateViewSet(viewsets.ModelViewSet):
    """Export 템플릿 관리"""
    serializer_class = ExportTemplateSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['format', 'template_type', 'is_public', 'is_default']

    def get_queryset(self):
        # 공개 템플릿 또는 본인이 만든 템플릿만 조회
        return ExportTemplate.objects.filter(
            is_active=True
        ).filter(
            models.Q(is_public=True) | models.Q(created_by=self.request.user)
        ).order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class ExportHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    """Export 이력 조회 (읽기 전용)"""
    serializer_class = ExportHistorySerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['project', 'format', 'status']

    def get_queryset(self):
        return ExportHistory.objects.filter(exported_by=self.request.user).order_by('-created_at')


class ReportScheduleViewSet(viewsets.ModelViewSet):
    """자동 보고서 스케줄 관리"""
    serializer_class = ReportScheduleSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['project', 'frequency', 'is_active']

    def get_queryset(self):
        return ReportSchedule.objects.filter(created_by=self.request.user).order_by('-created_at')

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class ExportDataViewSet(viewsets.ViewSet):
    """프로젝트 데이터 내보내기 - Excel / CSV / JSON"""
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=['get'], url_path='excel')
    def excel(self, request):
        """Excel 내보내기 (openpyxl 설치 시) / CSV fallback"""
        project_id = request.query_params.get('project')
        if not project_id:
            return Response({'error': '프로젝트 ID가 필요합니다.'}, status=400)

        try:
            from apps.projects.models import Project, Criteria
            project = Project.objects.get(id=project_id)
            criteria_qs = Criteria.objects.filter(
                project=project, is_active=True, type='criteria'
            ).order_by('level', 'order')
            alternatives_qs = Criteria.objects.filter(
                project=project, is_active=True, type='alternative'
            ).order_by('order')

            try:
                import openpyxl
                wb = openpyxl.Workbook()

                ws1 = wb.active
                ws1.title = '프로젝트 정보'
                ws1.append(['항목', '값'])
                ws1.append(['프로젝트명', project.title])
                ws1.append(['설명', project.description or ''])
                ws1.append(['상태', project.status])
                ws1.append(['생성일', str(project.created_at.date())])
                ws1.append(['기준 수', criteria_qs.count()])
                ws1.append(['대안 수', alternatives_qs.count()])

                ws2 = wb.create_sheet('평가 기준')
                ws2.append(['이름', '설명', '레벨', '순서', '가중치'])
                for c in criteria_qs:
                    ws2.append([c.name, c.description or '', c.level, c.order, c.weight or 0])

                ws3 = wb.create_sheet('대안')
                ws3.append(['이름', '설명', '순서'])
                for a in alternatives_qs:
                    ws3.append([a.name, a.description or '', a.order])

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                response = HttpResponse(
                    output.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                safe_title = project.title.replace(' ', '_')[:50]
                response['Content-Disposition'] = f'attachment; filename="{safe_title}.xlsx"'
                return response

            except ImportError:
                response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
                safe_title = project.title.replace(' ', '_')[:50]
                response['Content-Disposition'] = f'attachment; filename="{safe_title}.csv"'
                writer = csv.writer(response)
                writer.writerow(['=== 프로젝트 정보 ==='])
                writer.writerow(['프로젝트명', project.title])
                writer.writerow(['설명', project.description or ''])
                writer.writerow(['상태', project.status])
                writer.writerow([])
                writer.writerow(['=== 평가 기준 ==='])
                writer.writerow(['이름', '레벨', '순서', '가중치'])
                for c in criteria_qs:
                    writer.writerow([c.name, c.level, c.order, c.weight or 0])
                writer.writerow([])
                writer.writerow(['=== 대안 ==='])
                writer.writerow(['이름', '순서'])
                for a in alternatives_qs:
                    writer.writerow([a.name, a.order])
                return response

        except Exception as e:
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['get'], url_path='pdf')
    def pdf(self, request):
        """PDF 내보내기 (준비 중)"""
        return Response({
            'message': 'PDF 내보내기 기능은 준비 중입니다.',
            'available_formats': ['excel', 'report']
        }, status=status.HTTP_501_NOT_IMPLEMENTED)

    @action(detail=False, methods=['get'], url_path='report')
    def report(self, request):
        """JSON 보고서 내보내기"""
        project_id = request.query_params.get('project')
        if not project_id:
            return Response({'error': '프로젝트 ID가 필요합니다.'}, status=400)

        try:
            from apps.projects.models import Project, Criteria
            project = Project.objects.get(id=project_id)
            criteria = list(Criteria.objects.filter(
                project=project, is_active=True, type='criteria'
            ).values('id', 'name', 'description', 'level', 'order', 'weight'))
            alternatives = list(Criteria.objects.filter(
                project=project, is_active=True, type='alternative'
            ).values('id', 'name', 'description', 'order'))

            return Response({
                'project': {
                    'id': str(project.id),
                    'title': project.title,
                    'description': project.description or '',
                    'status': project.status,
                    'created_at': project.created_at.isoformat(),
                    'updated_at': project.updated_at.isoformat(),
                },
                'criteria': [
                    {**c, 'id': str(c['id'])} for c in criteria
                ],
                'alternatives': [
                    {**a, 'id': str(a['id'])} for a in alternatives
                ],
                'summary': {
                    'criteria_count': len(criteria),
                    'alternatives_count': len(alternatives),
                }
            })
        except Exception as e:
            return Response({'error': str(e)}, status=500)
